# backend/core/excel_handler.py
import os
import io
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from typing import Dict, List
import logging
from .constants import PosicoesExcel

class ExcelHandler:
    def __init__(self):
        self._cache_merged = {}
    
    def escrever_em_mesclada(self, ws, cell: str, valor):
        if ws not in self._cache_merged:
            self._cache_merged[ws] = list(ws.merged_cells.ranges)
        
        for merged_range in self._cache_merged[ws]:
            if cell in merged_range:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                ws.cell(row=min_row, column=min_col).value = valor
                return
        ws[cell] = valor
    
    async def criar_planilha(self, dados_por_fibra: List[Dict], modelo_content: bytes, 
                           nome_pasta: str, comprimento_bobina: float, 
                           operador: str) -> bytes:
        """Cria planilha Excel e retorna como bytes"""
        
        if not dados_por_fibra:
            raise ValueError("Nenhum dado para processar")
        
        primeiro = dados_por_fibra[0]
        
        # Carregar template do Excel a partir dos bytes
        wb = load_workbook(io.BytesIO(modelo_content))
        ws = wb.active
        
        data = primeiro.get("data", "")
        tipo = primeiro.get("tipo", "")
        distancia_final = primeiro.get("distancia_final", 0)
        metragem_total = int(distancia_final - comprimento_bobina)
        
        comentario = f"Comentários: Testes executados nas direções AB e BA, utilizando Bobina de Testes de {int(comprimento_bobina)} metros."
        
        # Preencher cabeçalho
        self.escrever_em_mesclada(ws, PosicoesExcel.BOBINA, f"BOBINA {nome_pasta}")
        self.escrever_em_mesclada(ws, PosicoesExcel.DATA, data)
        self.escrever_em_mesclada(ws, PosicoesExcel.METRAGEM_TOTAL, metragem_total)
        self.escrever_em_mesclada(ws, PosicoesExcel.LOTE, nome_pasta)
        self.escrever_em_mesclada(ws, PosicoesExcel.TIPO, tipo)
        self.escrever_em_mesclada(ws, PosicoesExcel.OPERADOR, operador)
        self.escrever_em_mesclada(ws, PosicoesExcel.COMENTARIO, comentario)
        
        # Preencher declives
        dados_por_fibra.sort(key=lambda x: x.get("fibra", 0))
        linha = PosicoesExcel.INICIO_DECLIVES
        for item in dados_por_fibra:
            declive = item.get("declive", 0)
            self.escrever_em_mesclada(ws, f"E{linha}", declive)
            linha += 1
        
        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()