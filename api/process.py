import os
import re
import zipfile
import tempfile
import shutil
import datetime
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse, HTMLResponse
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import pdfplumber

app = FastAPI()

def escrever_em_mesclada(ws, cell, valor):
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            ws.cell(row=min_row, column=min_col).value = valor
            return
    ws[cell] = valor

def extrair_dados_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        texto = ""
        for page in pdf.pages:
            texto += page.extract_text() + "\n"

    nome_arquivo = os.path.basename(pdf_path)

    bobina_match = re.search(r'B-(\d+)', nome_arquivo)
    bobina = bobina_match.group(1) if bobina_match else None

    data_match = re.search(r'Data de impressão\s*:\s*(\d{2}/\d{2}/\d{4})', texto)
    data = data_match.group(1) if data_match else ""

    fim_match = re.search(r'Fim da Fibra m\s+([\d\.]+)', texto)
    if fim_match:
        distancia_final = float(fim_match.group(1))
    else:
        distancias = re.findall(r'(\d+\.\d+)\s*m', texto)
        distancia_final = max([float(d) for d in distancias]) if distancias else 0

    declive_final = 0.0
    declive_match = re.search(r'Declive\s*\(?\s*dB/km\s*\)?\s*:\s*([-+]?\d+\.\d+)', texto, re.IGNORECASE)
    if declive_match:
        declive_final = float(declive_match.group(1))
    else:
        for linha in texto.split("\n"):
            if re.match(r'^\s*\d+\s', linha):
                numeros = re.findall(r'-?\d+\.\d+', linha)
                if len(numeros) >= 4:
                    declive_final = float(numeros[2])
                    break

    lote = ""
    tipo = ""
    ident = re.search(r'Ident\. Cabo.*?:\s*([A-Z0-9]+)\s+([A-Z0-9]+)', texto)
    if ident:
        lote = ident.group(1)
        tipo = ident.group(2)

    return {
        "bobina": bobina,
        "data": data,
        "distancia_final": distancia_final,
        "declive": declive_final,
        "lote": lote,
        "tipo": tipo
    }

def preencher_excel(dados_por_fibra, modelo_excel, pasta_saida, comprimento_bobina, operador):
    if not dados_por_fibra:
        return None

    primeiro = dados_por_fibra[0]
    bobina = primeiro["bobina"] or f"ERRO_{datetime.datetime.now().strftime('%H%M%S')}"
    nome_saida = f"BOBINA_{bobina}.xlsx"
    caminho_saida = os.path.join(pasta_saida, nome_saida)

    shutil.copyfile(modelo_excel, caminho_saida)
    wb = load_workbook(caminho_saida)
    ws = wb.active

    metragem_calculada = primeiro["distancia_final"] - comprimento_bobina
    metragem_total = max(0, round(metragem_calculada))

    escrever_em_mesclada(ws, "A8", f"BOBINA {bobina}")
    escrever_em_mesclada(ws, "E8", primeiro["data"])
    escrever_em_mesclada(ws, "G11", metragem_total)
    escrever_em_mesclada(ws, "G10", primeiro["lote"])
    escrever_em_mesclada(ws, "C9", primeiro["tipo"])
    escrever_em_mesclada(ws, "A67", operador)
    
    comentario = f"Comentários: Os Teste foram executados nas direções AB e BA, utilizando Bobina de Testes de {int(comprimento_bobina)} metros."
    escrever_em_mesclada(ws, "A63", comentario)

    linha = 15
    dados_por_fibra.sort(key=lambda x: x["fibra"])
    for item in dados_por_fibra:
        escrever_em_mesclada(ws, f"E{linha}", item["declive"])
        linha += 1

    wb.save(caminho_saida)
    return caminho_saida

@app.post("/api/process")
async def processar(
    zip_file: UploadFile = File(...),
    comprimento: float = Form(...),
    operador: str = Form(...)
):
    if not zip_file.filename.endswith('.zip'):
        raise HTTPException(400, "O arquivo enviado deve ser um ZIP.")

    temp_dir = tempfile.mkdtemp()
    extract_dir = os.path.join(temp_dir, "extraido")
    output_dir = os.path.join(temp_dir, "planilhas")
    modelo_path = None

    try:
        zip_path = os.path.join(temp_dir, "upload.zip")
        with open(zip_path, "wb") as f:
            f.write(await zip_file.read())

        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(extract_dir)

        for root, dirs, files in os.walk(extract_dir):
            for file in files:
                if file.endswith('.xlsx') and 'modelo' in file.lower():
                    modelo_path = os.path.join(root, file)
                    break
            if modelo_path:
                break

        if not modelo_path:
            raise HTTPException(400, "Nenhum arquivo modelo Excel foi encontrado no ZIP.")

        pre_lancamento_dir = None
        for root, dirs, files in os.walk(extract_dir):
            if os.path.basename(root) == "PRE-LANÇAMENTO":
                pre_lancamento_dir = root
                break

        if not pre_lancamento_dir:
            raise HTTPException(400, "O ZIP deve conter uma pasta chamada 'PRE-LANÇAMENTO'")

        pastas_bobinas = [p for p in os.listdir(pre_lancamento_dir) if os.path.isdir(os.path.join(pre_lancamento_dir, p))]
        if not pastas_bobinas:
            raise HTTPException(400, "Nenhuma subpasta (bobina) encontrada dentro de PRE-LANÇAMENTO.")

        os.makedirs(output_dir, exist_ok=True)

        for pasta in pastas_bobinas:
            caminho_pasta = os.path.join(pre_lancamento_dir, pasta)
            dados_fibras = []

            for arquivo in os.listdir(caminho_pasta):
                if arquivo.lower().endswith(".pdf"):
                    caminho_pdf = os.path.join(caminho_pasta, arquivo)
                    dados = extrair_dados_pdf(caminho_pdf)

                    fibra_match = re.search(r'F(\d+)', arquivo)
                    fibra = int(fibra_match.group(1)) if fibra_match else 0
                    dados["fibra"] = fibra
                    dados_fibras.append(dados)

            if dados_fibras:
                preencher_excel(dados_fibras, modelo_path, output_dir, comprimento, operador)

        zip_output = os.path.join(temp_dir, "resultado.zip")
        with zipfile.ZipFile(zip_output, 'w') as zf:
            for root, dirs, files in os.walk(output_dir):
                for file in files:
                    if file.endswith('.xlsx'):
                        full_path = os.path.join(root, file)
                        arcname = os.path.relpath(full_path, output_dir)
                        zf.write(full_path, arcname)

        def iterfile():
            with open(zip_output, 'rb') as f:
                yield from f

        return StreamingResponse(
            iterfile(),
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=planilhas_geradas.zip"}
        )

    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

@app.get("/")
async def root():
    index_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "index.html")
    if os.path.exists(index_path):
        with open(index_path, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse(content="<h1>Index.html não encontrado</h1>", status_code=404)